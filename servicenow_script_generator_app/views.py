from django.shortcuts import render, redirect, get_object_or_404
from .models import Category, SubCategory
from django.contrib import messages
from collections import defaultdict
from django.http import JsonResponse, HttpResponse
from openpyxl.utils import get_column_letter
import json, csv, openpyxl

def home(request):
  # GET request – just render everything
  return render(request, "home.html")

def view_cat_subcat(request):
  # GET request – just render everything
  categories = Category.objects.prefetch_related("subcategories").all()
  subcategories = SubCategory.objects.select_related("category").all()

  # Category search
  category_search = request.GET.get("category_search", "").strip()
  if category_search:
    categories = categories.filter(name__icontains=category_search)

  # Subcategory search
  sub_search = request.GET.get("subcategory_search", "").strip()
  # default: subcategory
  search_by = request.GET.get("search_by", "subcategory")

  if sub_search:
    if search_by == "category":
      subcategories = subcategories.filter(category__name__icontains=sub_search)
    # search by subcategory name
    else:  
      subcategories = subcategories.filter(name__icontains=sub_search)

  return render(request, "view_cat_subcat.html", {
    "categories": categories,
    "subcategories": subcategories,
    "category_count": categories.count(),
    "subcategory_count": subcategories.count(),
  })

# Helper function to normalize names
def normalize_name(name):
  #Normalize names for comparison (remove casing and spacing)
  if not name:
      return ""
  return "".join(name.lower().split())

def add_cat_subcat(request):
  if request.method == "POST":
    # Add all combos (JSON payload from hidden input)
    # Example payload:
    # {
    #   "category": "Networking",
    #   "subcategories": ["Routers", "Switches", "Cables"]
    # }
    all_combos_json = request.POST.get("all_combos")
    if all_combos_json:
      try:
        all_combos = json.loads(all_combos_json)

        # Build a lookup of normalized categories for fast access
        existing_categories = {
          normalize_name(cat.name): cat for cat in Category.objects.all()
        }

        # Build a lookup of (category_id, normalized_sub_name) normalized subcategories  for fast lookup
        existing_sub_pairs = {
          (sub.category_id, normalize_name(sub.name))
          for sub in SubCategory.objects.select_related("category").all()
        }
        
        # Make sure duplciates doesn't exist in SNOW tables and add if they are unique
        for combo in all_combos:
          category_name = combo.get("category", "").strip()
          subcategories = [
            s.strip() for s in combo.get("subcategories", []) if s.strip()
          ]

          normalized_category = normalize_name(category_name)

          # If normalized_category NOT in existing_categories, create it
          if normalized_category not in existing_categories:
            category = Category.objects.create(
              name=category_name,
              sequence=None   
            )
            existing_categories[normalized_category] = category
          else:
            category = existing_categories[normalized_category]

          # Add only unique (category, subcategory) pairs. 
          for sub_name in subcategories:
            normalized_sub = normalize_name(sub_name)
            pair = (category.id, normalized_sub)

            if pair not in existing_sub_pairs:
              SubCategory.objects.create(
                category=category,
                name=sub_name,
                sequence=None 
              )
              existing_sub_pairs.add(pair)

      except json.JSONDecodeError:
        print("Invalid JSON payload for all_combos.")

    # Redirect to add_to_snow page after processing
      return redirect("generate_scripts_page")

  # GET request – just render everything
  categories = Category.objects.prefetch_related("subcategories").all()
  return render(request, "add_cat_subcat.html", {"categories": categories})

# CRUD: Category
def edit_category(request, pk):
  category = get_object_or_404(Category, pk=pk)
  if request.method == "POST":
    new_name = request.POST.get("name", "").strip()
    if new_name:
      category.name = new_name
      category.save()
      return redirect("view_cat_subcat")
  return render(request, "edit.html", {
    "object": category,
    "type": "Category"
  })

def delete_category(request, pk):
  category = get_object_or_404(Category, pk=pk)
  subcategories_count = category.subcategories.count()

  if request.method == "POST":
    if subcategories_count == 0:
      category.delete()
      messages.success(request, f"Category '{category.name}' deleted successfully.")
      return redirect("view_cat_subcat")
    # Cannot delete category if it has subcategories
    else: 
      messages.error(request, f"Cannot delete category '{category.name}' because it has {subcategories_count} associated subcategories.")
      return redirect("view_cat_subcat")

  return render(request, "confirm_delete.html", {
    "object": category,
    "type": "Category",
    "cancel_url": "view_cat_subcat"
  })
    
def delete_all_categories(request):
  if request.method == "POST":
    Category.objects.all().delete()
    messages.success(request, "All categories have been deleted.")
  return redirect("view_cat_subcat")

# CRUD: SubCategory
def edit_subcategory(request, pk):
  sub = get_object_or_404(SubCategory, pk=pk)
  if request.method == "POST":
    new_name = request.POST.get("name", "").strip()
    if new_name:
      sub.name = new_name
      sub.save()
      return redirect("view_cat_subcat")
  return render(request, "edit.html", {
    "object": sub,
    "type": "Subcategory"
  })

def delete_subcategory(request, pk):
  sub = get_object_or_404(SubCategory, pk=pk)
  if request.method == "POST":
    sub.delete()
    return redirect("view_cat_subcat")
  return render(request, "confirm_delete.html", {
    "object": sub,
    "type": "Subcategory",
    "cancel_url": "view_cat_subcat"
  })

def delete_all_subcategories(request):
  if request.method == "POST":
    SubCategory.objects.all().delete()
    messages.success(request, "All subcategories have been deleted.")
  return redirect("view_cat_subcat")

def generate_scripts_page(request):
  return render(request, "generate_scripts.html")

def upload_category_csv(request):
  upload_message = None
  categories = Category.objects.all()

  if request.method == "POST" and request.FILES.get("category_csv_file"):
    csv_file = request.FILES["category_csv_file"]

    if not csv_file.name.endswith('.csv'):
      upload_message = "❌ Please upload a valid CSV file."
    else:
      # Read and decode the file
      decoded_file = csv_file.read().decode("utf-8").splitlines()
      reader = csv.DictReader(decoded_file)

      # Clear old categories
      Category.objects.all().delete()

      # Use a dict to prevent duplicates
      categories_dict = {}

      for row in reader:
        value = row.get("value").strip()
        inactive = row.get("inactive").lower()
        sequence_raw = row.get("sequence", "").strip()

        # skip blank and inactive rows 
        if value and inactive == "false":  
          # Normalize key
          normalized_key = "".join(value.lower().split())  # lower + remove spaces

          # Only add if not already present
          if normalized_key not in categories_dict:
            sequence = int(sequence_raw) if sequence_raw.isdigit() else None

            categories_dict[normalized_key] = {
              "name": value,
              "sequence": sequence,
            }

        # Convert to model instances
        categories_to_create = [
          Category(name=data["name"], sequence=data["sequence"])
          for data in categories_dict.values()
        ]

      Category.objects.bulk_create(categories_to_create)
      upload_message = f"✅ Successfully uploaded {len(categories_to_create)} categories from CSV."

      categories = Category.objects.all()

  return render(request, "home.html", {
    "categories": categories,
    "upload_message": upload_message
  })

def upload_subcategory_csv(request):
  upload_message = None
  subcategories = SubCategory.objects.select_related("category").all()

  if request.method == "POST" and request.FILES.get("subcategory_csv_file"):
    csv_file = request.FILES["subcategory_csv_file"]

    if not csv_file.name.endswith(".csv"):
      upload_message = "❌ Please upload a valid CSV file."
    else:
      # Read and decode the file
      decoded_file = csv_file.read().decode("utf-8").splitlines()
      reader = csv.DictReader(decoded_file)

      # Clear old subcategories
      SubCategory.objects.all().delete()

      # Prevent duplicates using normalized keys
      subcategory_dict = {}
      missing_categories = []

      for row in reader:
        value = row.get("value").strip()
        inactive = row.get("inactive", "").strip().lower()
        category_name = row.get("dependent_value").strip() or ""
        sequence_raw = row.get("sequence", "").strip()

        # Skip invalid rows
        if not value or inactive != "false":
          continue

        try:
          category = Category.objects.get(name=category_name)
        except Category.DoesNotExist:
          missing_categories.append(value)
          continue

        # Normalize key by combining category + subcategory
        normalized_key = (
          "".join(category_name.lower().split())
          + "|"
          + "".join(value.lower().split())
        )

        # Add only if unique for that category
        if normalized_key not in subcategory_dict:
          sequence = int(sequence_raw) if sequence_raw.isdigit() else None

          subcategory_dict[normalized_key] = {
            "category": category,
            "name": value,
            "sequence": sequence,
          }

      # Convert to model instances
      subcategories_to_create = [
        SubCategory(
          category=data["category"],
          name=data["name"],
          sequence=data["sequence"],
        )
        for data in subcategory_dict.values()
      ]

      SubCategory.objects.bulk_create(subcategories_to_create)

      # Build upload message
      success_count = len(subcategories_to_create)
      if missing_categories:
        upload_message = (
          f"✅ Uploaded {success_count} unique subcategories.\n"
          f"⚠️ Skipped {len(missing_categories)} subcategories because their category was missing/didn't match the Category table:\n"
          + ", ".join(missing_categories)
        )
      else:
        upload_message = f"✅ Successfully uploaded {success_count} unique subcategories from CSV."

      subcategories = SubCategory.objects.select_related("category").all()

  return render(
    request,
    "home.html",
    {
      "subcategories": subcategories,
      "upload_message_subcategories": upload_message,
    },
  )

def generate_category_excel(request):
  # Create workbook and sheet
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Categories"

  # Define column headers
  headers = [
    "Dependent value",
    "Element",
    "Hint",
    "Inactive",
    "Label",
    "Language",
    "Sequence",
    "Synonyms",
    "Value",
  ]

  # Write header row
  for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

  # Populate rows from Category model
  categories = Category.objects.all().order_by("sequence", "name")

  row_num = 2
  for cat in categories:
    ws.cell(row=row_num, column=1, value=None)          # Dependent value
    ws.cell(row=row_num, column=2, value="category")    # Element
    ws.cell(row=row_num, column=3, value=None)          # Hint
    ws.cell(row=row_num, column=4, value="FALSE")       # Inactive
    ws.cell(row=row_num, column=5, value=cat.name)      # Label
    ws.cell(row=row_num, column=6, value=None)          # Language
    ws.cell(row=row_num, column=7, value=cat.sequence)  # Sequence
    ws.cell(row=row_num, column=8, value=None)          # Synonyms
    ws.cell(row=row_num, column=9, value=cat.name)      # Value
    row_num += 1

  # Prepare HTTP response with Excel file
  response = HttpResponse(
    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
  )
  response["Content-Disposition"] = 'attachment; filename="categories.xlsx"'
  wb.save(response)
  return response

def generate_subcategory_excel(request):
  # Create workbook and sheet
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Subcategories"

  # Define column headers
  headers = [
    "Dependent value",
    "Element",
    "Hint",
    "Inactive",
    "Label",
    "Language",
    "Sequence",
    "Synonyms",
    "Value",
  ]

  # Write header row
  for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

  # Retrieve all subcategories, ordered cleanly
  subcategories = (
    SubCategory.objects
    .select_related("category")
    .order_by("category__sequence", "category__name", "sequence", "name")
  )

  row_num = 2
  for sub in subcategories:
    ws.cell(row=row_num, column=1, value=sub.category.name)  # Dependent value (parent Category)
    ws.cell(row=row_num, column=2, value="subcategory")       # Element
    ws.cell(row=row_num, column=3, value=None)                # Hint
    ws.cell(row=row_num, column=4, value="FALSE")             # Inactive
    ws.cell(row=row_num, column=5, value=sub.name)            # Label
    ws.cell(row=row_num, column=6, value=None)                # Language
    ws.cell(row=row_num, column=7, value=sub.sequence)        # Sequence
    ws.cell(row=row_num, column=8, value=None)                # Synonyms
    ws.cell(row=row_num, column=9, value=sub.name)            # Value
    row_num += 1

  # Prepare HTTP response with Excel file
  response = HttpResponse(
    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
  )
  response["Content-Disposition"] = 'attachment; filename="subcategories.xlsx"'
  wb.save(response)
  return response

def build_category_based_on_subcat_string(category_based_on_subcat, all_categories):
  js = []
  js.append("function onChange(isLoading) {")
  js.append("    if (isLoading) {")
  js.append("        return;")
  js.append("    }")
  js.append("")

  # Dump allCategories as JSON (safe for JS)
  js.append(f"  var allCategories = {json.dumps(all_categories)};")
  js.append("")
  js.append("  var subCat = g_form.getValue('subcategory');")
  js.append("  var category = g_form.getControl('category');")
  js.append("")
  js.append("  g_form.clearOptions('category');")
  js.append("")
  js.append("  switch (subCat) {")
  js.append('    case "":')
  js.append("      addOptions(category, allCategories);")
  js.append("      break;")
  js.append("")

  # Loop through mappings to create cases
  for sub, cats in category_based_on_subcat.items():
    if sub == '':
      continue
    case_line = f'    case "{sub}":'
    options = json.dumps([''] + cats)  
    js.append(case_line)
    js.append(f"      addOptions(category, {options});")
    js.append("      break;")
    js.append("")

  js.append("  }")
  js.append("")
  js.append("  function addOptions(element, options) {")
  js.append("    options.forEach(function(option) {")
  js.append("      var label = option === '' ? '-- None --' : option;")
  js.append("      g_form.addOption('category', option, label);")
  js.append("    });")
  js.append("  }")
  js.append("}")

  return "\n".join(js)

def build_subcategory_based_on_category_string(subcat_based_on_category, all_subcategories):
  js = []
  js.append("function onChange(isLoading) {")
  js.append("    if (isLoading) {")
  js.append("        return;")
  js.append("    }")
  js.append("")

  # Dump allCategories as JSON (safe for JS)
  js.append(f"  var allSubCategories = {json.dumps(all_subcategories)};")
  js.append("")
  js.append("  var category = g_form.getValue('category');")
  js.append("  var subCat = g_form.getControl('subcategory');")
  js.append("")
  js.append("  g_form.clearOptions('subcategory');")
  js.append("")
  js.append("  switch (category) {")
  js.append('    case "":')
  js.append("      addOptions(subCat, allSubCategories);")
  js.append("      break;")
  js.append("")

  # Loop through mappings to create cases
  for cat, subs in subcat_based_on_category.items():
    if cat == '':
      continue
    case_line = f'    case "{cat}":'
    options = json.dumps([''] + subs)  
    js.append(case_line)
    js.append(f"      addOptions(subcategory, {options});")
    js.append("      break;")
    js.append("")
    

  js.append("  }")
  js.append("")
  js.append("  function addOptions(element, options) {")
  js.append("    options.forEach(function(option) {")
  js.append("      var label = option === '' ? '-- None --' : option;")
  js.append("      g_form.addOption('subcategory', option, label);")
  js.append("    });")
  js.append("  }")
  js.append("}")

  return "\n".join(js)

def generate_scripts(request):
  categories = Category.objects.all()
  subcategories = SubCategory.objects.select_related("category").all()

  # List of all categories and subcategories 
  all_categories = [''] + [cat.name for cat in categories]
  all_subcategories = [''] + [sub.name for sub in subcategories]
  
  # Create defaultdict to hold category/subcategory mappings
  subcat_based_on_category = defaultdict(list)
  category_based_on_subcat = defaultdict(list)

  # Map subcategories to their categories and vice versa
  for sub in subcategories:
    subcat_based_on_category[sub.category.name].append(sub.name)
    category_based_on_subcat[sub.name].append(sub.category.name)

  # Build JS strings
  category_based_on_subcat_string = build_category_based_on_subcat_string(category_based_on_subcat, all_categories)
  subcat_based_on_category_string = build_subcategory_based_on_category_string(subcat_based_on_category, all_subcategories)

  return JsonResponse({
    "category_based_on_subcat_script": category_based_on_subcat_string,
    "subcat_based_on_category_script": subcat_based_on_category_string,
  })